import os
import openpyxl
import pandas as pd
import plotly.express as px

from .constants import Indicateurs, Impurete, Ferrite, ONO, THIELMANN, MAYER, Quality_ref

# Utilisez ces constantes dans votre code


def remove_outliers(df, elements):
    """
    Supprime les valeurs extrêmes d'un DataFrame.

    Args:
    - df : DataFrame contenant les données
    - elements : Liste des noms de colonnes sur lesquelles rechercher les valeurs aberrantes

    Returns:
    - df_cleaned : DataFrame nettoyé, sans les valeurs aberrantes
    - outliers_df : DataFrame contenant les valeurs extrêmes supprimées
    """
    # Copier le DataFrame pour éviter les modifications inattendues
    df = df.copy()
    outliers_df = pd.DataFrame()  # DataFrame pour stocker les lignes supprimées

    while True:
        has_outliers = False  # Indicateur pour vérifier s'il y a des outliers

        for element in elements:
            # Calculer Q1, Q3 et l'IQR
            Q1 = df[element].quantile(0.25)
            Q3 = df[element].quantile(0.75)
            IQR = Q3 - Q1

            # Calculer Q1 - 1.5 * IQR et Q3 + 1.5 * IQR
            lower_bound = Q1 - 1.5 * IQR
            upper_bound = Q3 + 1.5 * IQR

            # Sélectionner les lignes qui sont en dehors de l'intervalle [lower_bound, upper_bound]
            outliers = df.loc[(df[element] < lower_bound) | (df[element] > upper_bound)]

            # Si des outliers sont trouvés, les stocker dans le DataFrame des outliers
            if not outliers.empty:
                has_outliers = True
                outliers_df = pd.concat([outliers_df, outliers], axis=0)
                # Supprimer les lignes correspondantes du DataFrame original
                df = df.drop(outliers.index)

        # Sortir de la boucle while si aucun outlier n'est trouvé
        if not has_outliers:
            break

    return df, outliers_df

def export_outliers_and_cleaned_data(df, recipe_name, output_dir):
    """
    Exporte les données avec et sans valeurs aberrantes dans un classeur Excel.

    Args:
    - df : DataFrame contenant les données
    - recipe_name : Nom de la recette
    - output_dir : Répertoire de sortie pour le fichier Excel

    Returns:
    - df : DataFrame nettoyé sans les valeurs aberrantes
    """
    # Suppression des outliers
    df, outliers = remove_outliers(df, Indicateurs)
    
    # Exportation des données conformes avec et sans outliers
    excel_path = os.path.join(output_dir, f'{recipe_name}.xlsx')

    # Charger le classeur Excel
    workbook = openpyxl.load_workbook(excel_path)

    # Exportation des données conformes avec et sans outliers
    with pd.ExcelWriter(excel_path, mode='a', engine='openpyxl',if_sheet_exists='replace') as writer:
        # Vérifier si les feuilles existent
        if 'Conforme_with_outliers' not in workbook.sheetnames :
            outliers.to_excel(writer, sheet_name=f'Conforme_with_outliers', index=False)
        if 'Conforme_without_outliers' not in workbook.sheetnames:
            df.to_excel(writer, sheet_name=f'Conforme_without_outliers', index=False)
    return df

def plot_and_save(df, y_column, save=None):
    """
    Plot and save graphs with and without outliers.

    Args:
    - df: DataFrame containing the data.
    - y_column: Name of the column containing the y-axis data.
    - output_dir: Path to the directory where the graphs will be saved.

    Returns:
    - df: DataFrame with outliers removed.
    - outliers: DataFrame containing the removed outliers.
    """



    # Graphe avec les valeurs extrêmes
    fig1 = px.scatter(df, x=df.index, y=y_column, title=f'{y_column} avec ses valeurs extrêmes')

    # Définir la taille de la figure
    # fig1.update_layout(width=665, height=500)
    # fig1.update_layout(width=485, height=500)
    fig1.update_layout(width=475, height=500)

    fig1.show()
    if save: 
        # Enregistrement de l'image
        recipe_name = df.iloc[0]['Recette']
        # Chemin vers le dossier
        folder_path = os.path.join('..', 'Images', recipe_name)
        # Vérifier si le dossier existe, sinon le créer
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        path_extreme = os.path.join(folder_path,f'{y_column}_avec_extreme.pdf')
        # fig1.write_image(path_extreme, format='pdf', scale=3)
        fig1.write_image(path_extreme, format='pdf', scale=2)

    # Suppression des outliers
    df, outliers = remove_outliers(df, Indicateurs)


    
    # Graphe sans les valeurs extrêmes
    fig2 = px.scatter(df, x=df.index, y=y_column, title=f'{y_column} sans ses valeurs extrêmes')
    
    # Définir la taille de la figure
    # fig2.update_layout(width=665, height=500)
    # fig2.update_layout(width=485, height=500)
    fig2.update_layout(width=475, height=500)

    fig2.show()
    if save: 
        path_no_extreme = os.path.join(folder_path, f'{y_column}_sans_extreme.pdf')
        # fig2.write_image(path_no_extreme, format='pdf', scale=3)
        fig2.write_image(path_no_extreme, format='pdf', scale=2)

    return 
