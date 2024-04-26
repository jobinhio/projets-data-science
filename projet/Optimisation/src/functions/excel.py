from openpyxl.utils.dataframe import dataframe_to_rows
import os
from openpyxl import Workbook


def write_dataframe_to_excel(df, dossier_data, new_sheet_name='Recette'):
    """
    Écrit un DataFrame dans une nouvelle feuille Excel ou écrase une feuille existante avec les données du DataFrame.

    Args:
        df (DataFrame): Le DataFrame à écrire dans le fichier Excel.
        excel_file_path (str): Chemin du fichier Excel existant.
        new_sheet_name (str, optional): Nom de la nouvelle feuille Excel. Par défaut, 'Recette'.

    Returns:
        None
    """


    # Créer le chemin complet du nouveau fichier Excel
    fichier_erreurs = os.path.join(dossier_data, 'erreurs.txt')

    # Vérifier si le fichier Excel existe déjà
    if os.path.exists(fichier_erreurs):
        # Supprimer le fichier existant
        os.remove(fichier_erreurs)

    workbook = Workbook()

    # Supprimer la première feuille s'il y en a une
    if workbook.sheetnames:
        first_sheet = workbook.sheetnames[0]
        workbook.remove(workbook[first_sheet])


    # Vérifier si la feuille existe déjà dans le classeur Excel
    if new_sheet_name in workbook.sheetnames:
        # Si la feuille existe déjà, effacer les données
        workbook.remove(workbook[new_sheet_name])

    # Ajouter la nouvelle feuille au classeur Excel
    workbook.create_sheet(title=new_sheet_name)

    # Charger la feuille nouvellement créée
    new_sheet = workbook[new_sheet_name]

    # Écrire les données dans la feuille
    for row in dataframe_to_rows(df, index=False, header=True):
        new_sheet.append(row)

    # Créer le chemin complet du nouveau fichier Excel
    nouveau_fichier_excel = os.path.join(dossier_data, 'Resultats.xlsx')

    # Sauvegarder le classeur Excel dans le nouveau fichier Excel
    workbook.save(nouveau_fichier_excel)

    # Fermer le classeur Excel
    workbook.close()