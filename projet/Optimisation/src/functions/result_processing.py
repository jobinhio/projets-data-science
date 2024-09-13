import pandas as pd
import numpy as np
import os
from .constants import Impurete_Values, ONO_Values
from .constants import THIELMANN_Values, MAYER_Values, Ferrite_Values
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import gc


def remove_old_recipes(dossier_data):
    """
    Supprime les anciens fichiers de recettes et de fichiers d'erreurs, s'ils existent, dans le dossier de données.

    Args:
    dossier_data (str): Chemin vers le dossier contenant les données.

    Returns:
    None
    """
    gc.collect()
    # Vérifier si le fichier résultats existe et le supprime
    fichier_recipes = os.path.join(dossier_data, 'recipes.xlsx')
    if os.path.exists(fichier_recipes):
        os.remove(fichier_recipes)
    
    # Vérifier si le fichier erreurs existe et le supprime
    fichier_erreurs = os.path.join(dossier_data, 'erreurs.txt')
    if os.path.exists(fichier_erreurs):
        os.remove(fichier_erreurs)

    # Vérifier si le fichier résultats existe et le supprime
    fichier_recipes = os.path.join(dossier_data, 'InputsOutputs','resultats.xlsx')
    if os.path.exists(fichier_recipes):
        os.remove(fichier_recipes)

    # Vérifier si le fichier résultats existe et le supprime
    fichier_recipes = os.path.join(dossier_data, 'InputsOutputs', 'resultatsComposition.xlsx')
    if os.path.exists(fichier_recipes):
        os.remove(fichier_recipes)

    # Vérifier si le fichier erreurs existe et le supprime
    fichier_erreurs = os.path.join(dossier_data, 'InputsOutputs','erreurs.txt')
    if os.path.exists(fichier_erreurs):
        os.remove(fichier_erreurs)
    return 


def construct_result_dataframe(res, df_mp_dispo, table_mp, constraints):
    """
    Construit un DataFrame contenant les résultats de l'optimisation ainsi que les contraintes à respecter.

    Args:
    res: Résultat de l'optimisation.
    df_mp_dispo (DataFrame): DataFrame contenant les données sur les matières premières disponibles.
    table_mp (DataFrame): DataFrame contenant les données sur les matières premières et leurs éléments.
    constraints (dict): Dictionnaire contenant les contraintes du problème.

    Returns:
    DataFrame: DataFrame contenant les résultats de l'optimisation.
    dict: Dictionnaire contenant les contraintes du résultat.
    """
    # Constantes pour les seuils Métalliques
    SEUIL_0 = 1e-20
    SEUIL_1 = 1e-20 # 1e-20 0.01

    # Création du DataFrame df_res
    df_res = df_mp_dispo[['Code Article','Article', 'Prix', 'Métallique ?']].copy()

    # Ajout de la colonne 'Proportion' avec les valeurs de res.x
    # df_res['Proportion'] = res.x
    df_res['Proportion'] = res.x[:df_mp_dispo[['Prix']].shape[0]]

    # Calcul de la colonne 'Valeur (/T)' en multipliant 'Proportion' par 'Prix'
    df_res['Valeur (/T)'] = df_res['Proportion'] * df_res['Prix']

    # Filtre des lignes avec des proportions non nulles seulement
    df_res = df_res[df_res['Proportion'] > 0]

    # Tri du DataFrame par 'Proportion' de manière décroissante
    df_res.sort_values(by='Proportion', ascending=False, inplace=True)

    # Initialisation des colonnes 'ONO' et 'Impurete' avec des valeurs NaN
    df_res['ONO'] = np.nan
    df_res['Impurete'] = np.nan

    # Filtrage des lignes en fonction de la valeur de la colonne 'Métallique ?' et du seuil correspondant
    df_res = df_res.loc[((df_res['Métallique ?'] == 0) & (df_res['Proportion'] >= SEUIL_0)) |
                        ((df_res['Métallique ?'] == 1) & (df_res['Proportion'] >= SEUIL_1))]

    # Suppression de la colonne 'Métallique ?' qui n'est plus nécessaire
    df_res.drop(columns=['Métallique ?'], inplace=True)

    # Fusion avec les éléments chimiques correspondant aux articles sélectionnés
    articles_selectionnes = df_res['Article'].tolist()
    elements_selectionnes = table_mp[table_mp['Article'].isin(articles_selectionnes)]
    df_res = pd.merge(df_res, elements_selectionnes, on=['Code Article','Article'], how='inner')

    # Ajout de la ligne de résultats
    df_res.loc[df_res.shape[0] , :] = np.nan
    df_res.loc[df_res.shape[0] , 'Article'] = 'Résultats'

    df_res.loc[df_res.shape[0] - 1, ['Proportion', 'Valeur (/T)']] = [df_res['Proportion'].sum(), df_res['Valeur (/T)'].sum()]

    # Calcul des proportions des éléments dans la fonte
    cols_elements = table_mp.columns[2:]
    proportions_elements = df_res[cols_elements].mul(df_res['Proportion'], axis=0).sum()

    # Ajout des valeurs des proportions des éléments
    df_res.loc[df_res.shape[0]-1, cols_elements] = proportions_elements 

    # Calcul des valeurs des indicateurs qualité
    impurete_values, ono_values = np.array(list(Impurete_Values.values())), np.array(list(ONO_Values.values()))
    df_res.loc[df_res.shape[0]-1, 'Impurete'] = impurete_values @ proportions_elements
    df_res.loc[df_res.shape[0]-1, 'ONO'] = ono_values @ proportions_elements

    contraints_res = {}
    contraints_res = proportions_elements.to_dict()
    contraints_res['Impurete'] = impurete_values @ proportions_elements
    contraints_res['ONO'] = ono_values @ proportions_elements
    contraints_res['Proportion_Total'] = df_res.loc[df_res.shape[0] - 1, ['Proportion']].values[0]

    # Ajout du recipes des contraintes de MP
    b_eq = constraints['b_eq'] 
    for key, value in b_eq.items() :
        for article in df_res['Article'] :
            if article == key :
                contraints_res[key] = df_res.loc[df_res['Article'] == key,'Proportion'].values[0]
    return df_res,contraints_res


def export_result(df_res1, df_res2, dossier_data, new_sheet_name):
    """
    Exporte deux DataFrames dans un fichier Excel, créant de nouvelles feuilles ou mettant à jour des feuilles existantes.

    Args:
    df_res1 (DataFrame): Le premier DataFrame à exporter (peut être None).
    df_res2 (DataFrame): Le deuxième DataFrame à exporter (peut être None).
    dossier_data (str): Chemin vers le dossier contenant les données.
    new_sheet_name (str): Nom de la nouvelle feuille Excel.

    Returns:
    None
    """
    sheet_name1, sheet_name2 = new_sheet_name + " Version 1", new_sheet_name + " Version 2"

    # Créer le chemin complet du fichier Excel
    fichier_recipes = os.path.join(dossier_data, 'recipes.xlsx')

    # Vérifier si le fichier Excel existe
    if os.path.exists(fichier_recipes):
        # Charger le classeur Excel existant
        workbook = load_workbook(fichier_recipes)
    else:
        # Créer un nouveau classeur s'il n'existe pas
        workbook = Workbook()

    # Ajouter ou mettre à jour la première feuille (si df_res1 n'est pas None)
    if df_res1 is not None:
        if sheet_name1 in workbook.sheetnames:
            feuille1 = workbook[sheet_name1]
            workbook.remove(feuille1)
            feuille1 = workbook.create_sheet(title=sheet_name1)
        else:
            feuille1 = workbook.create_sheet(title=sheet_name1)

        # Écrire le premier DataFrame dans la première feuille
        for r_idx, row in enumerate(dataframe_to_rows(df_res1, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                feuille1.cell(row=r_idx, column=c_idx, value=value)

    # Ajouter ou mettre à jour la deuxième feuille (si df_res2 n'est pas None)
    if df_res2 is not None:
        if sheet_name2 in workbook.sheetnames:
            feuille2 = workbook[sheet_name2]
            workbook.remove(feuille2)
            feuille2 = workbook.create_sheet(title=sheet_name2)
        else:
            feuille2 = workbook.create_sheet(title=sheet_name2)

        # Écrire le deuxième DataFrame dans la deuxième feuille
        for r_idx, row in enumerate(dataframe_to_rows(df_res2, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                feuille2.cell(row=r_idx, column=c_idx, value=value)

    # Supprimer la feuille par défaut si elle existe et n'est pas utilisée
    if 'Sheet' in workbook.sheetnames :
        default_sheet = workbook['Sheet']
        workbook.remove(default_sheet)

    # Sauvegarder le classeur
    workbook.save(fichier_recipes)
    workbook.close()
    gc.collect()

    return


def save_errors(erreurs, dossier_data,recette):
    """
    Sauvegarde les erreurs dans un fichier texte.

    Args:
    erreurs (list): Liste des erreurs à sauvegarder.
    dossier_data (str): Chemin vers le dossier contenant les données.
    recette (str): Nom de la recette associée aux erreurs.

    Returns:
    None
    """
    # Sauvegarder les erreurs dans un fichier texte
    # fichier_erreurs = os.path.join(dossier_data, 'erreurs.txt')
    fichier_erreurs = os.path.join(dossier_data, 'InputsOutputs', 'erreurs.txt')
    with open(fichier_erreurs, "a+", encoding="utf-8") as f:
        message = "Pour la recette " + recette + ":"
        f.write(message+ "\n")
        for erreur in erreurs:
            f.write(erreur + "\n")
        f.write("\n")
    return 

def gestion_resultats(erreurs1, res1, erreurs2, res2, df_mp_dispo, table_mp, constraints, dossier_data, recette):
    if not erreurs1 and not erreurs2:
        # Aucun des deux cas n'a d'erreurs : les deux solutions sont valides
        df_res1, contraints_res1 = construct_result_dataframe(res1, df_mp_dispo, table_mp, constraints)
        df_res2, contraints_res2 = construct_result_dataframe(res2, df_mp_dispo, table_mp, constraints)
        export_result(df_res1, df_res2, dossier_data, recette)
        print(f"Le problème pour la recette {recette} admet une solution pour les deux formulations.")
    
    elif not erreurs1:
        # Seule la première formulation est valide
        df_res1, contraints_res1 = construct_result_dataframe(res1, df_mp_dispo, table_mp, constraints)
        df_res2 = None  # Pas de solution pour res2
        export_result(df_res1, df_res2, dossier_data, recette)
        print(f"Le problème pour la recette {recette} admet une solution uniquement pour la première formulation.")
        save_errors(erreurs2, dossier_data, recette)
    
    elif not erreurs2:
        # Seule la deuxième formulation est valide
        df_res2, contraints_res2 = construct_result_dataframe(res2, df_mp_dispo, table_mp, constraints)
        df_res1 = None  # Pas de solution pour res1
        export_result(df_res1, df_res2, dossier_data, recette)
        print(f"Le problème pour la recette {recette} admet une solution uniquement pour la deuxième formulation.")
        save_errors(erreurs1, dossier_data, recette)
    
    else:
        # Les deux formulations ont des erreurs
        save_errors(erreurs1, dossier_data, recette)
        save_errors(erreurs2, dossier_data, recette)
        print(f"Les erreurs des deux formulations pour la recette {recette} ont été enregistrées dans un fichier.")
    
    return



#----------------- FDN -----------------


def construct_FDNresult_dataframe(res, df_mp_dispo, table_mp):
    """
    Construit un DataFrame contenant les résultats de l'optimisation ainsi que les contraintes à respecter.

    Args:
    res: Résultat de l'optimisation.
    df_mp_dispo (DataFrame): DataFrame contenant les données sur les matières premières disponibles.
    table_mp (DataFrame): DataFrame contenant les données sur les matières premières et leurs éléments.
    constraints (dict): Dictionnaire contenant les contraintes du problème.

    Returns:
    DataFrame: DataFrame contenant les résultats de l'optimisation.
    dict: Dictionnaire contenant les contraintes du résultat.
    """
    # Constantes pour les seuils Métalliques
    SEUIL_0 = 1e-20
    SEUIL_1 = 1e-20 # 1e-20 0.01

    # Création du DataFrame df_res
    df_res = df_mp_dispo[['Article', 'Métallique ?']].copy()

    # Ajout de la colonne 'Proportion' avec les valeurs de res.x
    # df_res['Proportion'] = res.x
    df_res['Proportion'] = res.x[:df_mp_dispo[['Prix']].shape[0]]

    # Calcul de la colonne 'Cout' en multipliant 'Proportion' par 'Prix'
    df_res['Cout'] = df_res['Proportion'] * df_mp_dispo['Prix']

    # Filtre des lignes avec des proportions non nulles seulement
    df_res = df_res[df_res['Proportion'] > 0]

    # Tri du DataFrame par 'Proportion' de manière décroissante
    df_res.sort_values(by='Proportion', ascending=False, inplace=True)


    # Filtrage des lignes en fonction de la valeur de la colonne 'Métallique ?' et du seuil correspondant
    df_res = df_res.loc[((df_res['Métallique ?'] == 0) & (df_res['Proportion'] >= SEUIL_0)) |
                        ((df_res['Métallique ?'] == 1) & (df_res['Proportion'] >= SEUIL_1))]

    # Suppression de la colonne 'Métallique ?' qui n'est plus nécessaire
    df_res.drop(columns=['Métallique ?'], inplace=True)


    # Pour le deuxieme fichier excel
    # Calcul des proportions des éléments dans la fonte
    cols_elements = table_mp.columns[2:].tolist()
    proportions_elements = table_mp[cols_elements].mul(df_res['Proportion'], axis=0).sum()

    # creer un nouveau dataframme df_resElements
    df_resElements = pd.DataFrame([proportions_elements], columns=cols_elements)


    # Calcul des valeurs des indicateurs qualité
    impurete_values, ono_values = np.array(list(Impurete_Values.values())), np.array(list(ONO_Values.values()))
    thielmann_values, mayer_values = np.array(list(THIELMANN_Values.values())), np.array(list(MAYER_Values.values()))
    ferrite_values = np.array(list(Ferrite_Values.values()))
    df_resElements['Impurete'] = impurete_values @ proportions_elements
    df_resElements['ONO'] = ono_values @ proportions_elements
    df_resElements['THIELMANN'] = thielmann_values @ proportions_elements
    df_resElements['MAYER'] = mayer_values @ proportions_elements
    df_resElements['Ferrite'] = 92.3 + ferrite_values @ proportions_elements


    # Cette partie est remplie uniquement par Salma
    df_resElements['Rm'] = np.nan
    df_resElements['Moyenne allongement'] = np.nan


    return df_res, df_resElements


def export_FDNresult(df_res1, df_res2, df_resElements1, df_resElements2, dossier_data, new_sheet_name):
    """
    Exporte deux DataFrames dans un fichier Excel, créant de nouvelles feuilles ou mettant à jour des feuilles existantes.

    Args:
    df_res (DataFrame): Le premier DataFrame à exporter (peut être None).
    df_resElements (DataFrame): Le deuxième DataFrame à exporter (peut être None).
    dossier_data (str): Chemin vers le dossier contenant les données.
    new_sheet_name (str): Nom de la nouvelle feuille Excel.

    Returns:
    None
    """
    sheet_name1, sheet_name2 = new_sheet_name + " Version 1", new_sheet_name + " Version 2"

    # Créer le chemin complet du fichier Excel
    fichier_resultats = os.path.join(dossier_data, 'InputsOutputs','resultats.xlsx')

    # Vérifier si le fichier Excel existe
    if os.path.exists(fichier_resultats):
        # Charger le classeur Excel existant
        workbook = load_workbook(fichier_resultats)
    else:
        # Créer un nouveau classeur s'il n'existe pas
        workbook = Workbook()

    # Ajouter ou mettre à jour la première feuille (si df_res n'est pas None)
    if df_res1 is not None:
        if sheet_name1 in workbook.sheetnames:
            feuille1 = workbook[sheet_name1]
            workbook.remove(feuille1)
            feuille1 = workbook.create_sheet(title=sheet_name1)
        else:
            feuille1 = workbook.create_sheet(title=sheet_name1)

        # Écrire le premier DataFrame dans la première feuille
        for r_idx, row in enumerate(dataframe_to_rows(df_res1, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                feuille1.cell(row=r_idx, column=c_idx, value=value)

    # Ajouter ou mettre à jour la deuxième feuille (si df_resElements n'est pas None)
    if df_res2 is not None:
        if sheet_name2 in workbook.sheetnames:
            feuille2 = workbook[sheet_name2]
            workbook.remove(feuille2)
            feuille2 = workbook.create_sheet(title=sheet_name2)
        else:
            feuille2 = workbook.create_sheet(title=sheet_name2)

        # Écrire le deuxième DataFrame dans la deuxième feuille
        for r_idx, row in enumerate(dataframe_to_rows(df_res2, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                feuille2.cell(row=r_idx, column=c_idx, value=value)

    # Supprimer la feuille par défaut si elle existe et n'est pas utilisée
    if 'Sheet' in workbook.sheetnames :
        default_sheet = workbook['Sheet']
        workbook.remove(default_sheet)

    # Sauvegarder le classeur
    workbook.save(fichier_resultats)
    workbook.close()
    # gc.collect()


    # Créer le chemin complet du fichier Excel
    fichier_resultatsComposition = os.path.join(dossier_data, 'InputsOutputs', 'resultatsComposition.xlsx')

    # Vérifier si le fichier Excel existe
    if os.path.exists(fichier_resultatsComposition):
        # Charger le classeur Excel existant
        workbook = load_workbook(fichier_resultatsComposition)
    else:
        # Créer un nouveau classeur s'il n'existe pas
        workbook = Workbook()

    # Ajouter ou mettre à jour la première feuille (si df_res n'est pas None)
    if df_resElements1 is not None:
        if sheet_name1 in workbook.sheetnames:
            feuille1 = workbook[sheet_name1]
            workbook.remove(feuille1)
            feuille1 = workbook.create_sheet(title=sheet_name1)
        else:
            feuille1 = workbook.create_sheet(title=sheet_name1)

        # Écrire le premier DataFrame dans la première feuille
        for r_idx, row in enumerate(dataframe_to_rows(df_resElements1, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                feuille1.cell(row=r_idx, column=c_idx, value=value)

    # Ajouter ou mettre à jour la deuxième feuille (si df_resElements n'est pas None)
    if df_resElements2 is not None:
        if sheet_name2 in workbook.sheetnames:
            feuille2 = workbook[sheet_name2]
            workbook.remove(feuille2)
            feuille2 = workbook.create_sheet(title=sheet_name2)
        else:
            feuille2 = workbook.create_sheet(title=sheet_name2)

        # Écrire le deuxième DataFrame dans la deuxième feuille
        for r_idx, row in enumerate(dataframe_to_rows(df_resElements2, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                feuille2.cell(row=r_idx, column=c_idx, value=value)

        # Supprimer la feuille par défaut si elle existe et n'est pas utilisée
    if 'Sheet' in workbook.sheetnames :
        default_sheet = workbook['Sheet']
        workbook.remove(default_sheet)

    # Sauvegarder le classeur
    workbook.save(fichier_resultatsComposition)
    workbook.close()
    gc.collect()

    return


def gestion_FDNresultats(erreurs1, res1, erreurs2, res2, df_mp_dispo, table_mp, constraints, dossier_data, recette):
    if not erreurs1 and not erreurs2:
        # Aucun des deux cas n'a d'erreurs : les deux solutions sont valides
        df_res1, df_resElements1 = construct_FDNresult_dataframe(res1, df_mp_dispo, table_mp)
        df_res2, df_resElements2 = construct_FDNresult_dataframe(res2, df_mp_dispo, table_mp)
        export_FDNresult(df_res1, df_res2, df_resElements1, df_resElements2, dossier_data, recette)
        print(f"Le problème pour la recette {recette} admet une solution pour les deux formulations.")
    
    elif not erreurs1:
        # Seule la première formulation est valide
        df_res1, df_resElements1 = construct_FDNresult_dataframe(res1, df_mp_dispo, table_mp)
        df_res2, df_resElements2 = None, None  # Pas de solution pour res2
        export_FDNresult(df_res1, df_res2, df_resElements1, df_resElements2, dossier_data, recette)
        print(f"Le problème pour la recette {recette} admet une solution uniquement pour la première formulation.")
        save_errors(erreurs2, dossier_data, recette)
    
    elif not erreurs2:
        # Seule la deuxième formulation est valide
        df_res2, df_resElements2 = construct_FDNresult_dataframe(res2, df_mp_dispo, table_mp)
        df_res1, df_resElements1 = None, None # Pas de solution pour res1
        export_FDNresult(df_res1, df_res2, df_resElements1, df_resElements2, dossier_data, recette)
        print(f"Le problème pour la recette {recette} admet une solution uniquement pour la deuxième formulation.")
        save_errors(erreurs1, dossier_data, recette)
    
    else:
        # Les deux formulations ont des erreurs
        save_errors(erreurs1, dossier_data, recette)
        save_errors(erreurs2, dossier_data, recette)
        print(f"Les erreurs des deux formulations pour la recette {recette} ont été enregistrées dans un fichier.")
    
    return

