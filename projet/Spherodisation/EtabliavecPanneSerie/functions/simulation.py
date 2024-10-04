import time
from .file_io import FileManager
from .plotting import PlotManager
from .utils import calcul_temps_limite, calcul_longueur_fil


from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import gc
import pandas as pd
import os
from pathlib import Path


class Simulation:
    def __init__(self, param_file, button_file, disa_file):
        # FileManager s'occupe de la lecture et de la gestion des fichiers
        self.file_manager = FileManager(param_file, button_file, disa_file)

        # Initialisation du gestionnaire de plots
        self.plot_manager = PlotManager()

        # Variables de simulation
        self.Mg = 0.045  # Quantité initiale de magnésium
        self.PFC = 3500  # Poids de fonte coulée initial
        self.t = 0  # Temps de simulation
        self.timedata = []
        self.Mgdata = []
        self.PFCdata = []

    def initialize_simulation(self):
        """Initialise la simulation en lisant les fichiers nécessaires via FileManager."""
        self.file_manager.read_parametres_generaux()
        self.file_manager.read_programme_disa()
        self.file_manager.read_file_boutons_et_parametres()

        if self.file_manager.parametres_generaux  and self.file_manager.ProgrammeDISA and self.file_manager.BoutonsEtParamètres:
            print("Simulation initialisée avec succès.")
        else:
            print("Échec de l'initialisation de la simulation. Vérifiez les fichiers.")


    def construct_result_dataframe(self):
        """
        Couleurs
        0=bleu
        1=Vert
        2=orange
        3=rouge

        """
        Mgmin = self.file_manager.parametres_generaux['Mgmin']
        PFCmax = self.file_manager.parametres_generaux['PFCmax']
        PPT = self.file_manager.BoutonsEtParamètres['PPT']

        colorMg = 1
        colorPFC = 0

        if (self.PFC + PPT) >= PFCmax and self.Mg <= Mgmin:
            colorMg = 3
        elif (self.PFC + PPT) <= PFCmax and self.Mg <= Mgmin:
            colorMg = 2
        else:
            colorMg = 1


        data = {
                'Courbe': ['MG', 'PFC'],
                'Temps(s)': [self.t, self.t],
                'ValY': [self.Mg, self.PFC],
                'Couleur': [colorMg, colorPFC]
            }
        
        df = pd.DataFrame(data)

        return df
   

    def export_result(self, df, dossier_data):
        """
        Exporter le DataFrame df dans un fichier Excel nommé 'sortie.xlsx' dans le dossier spécifié.
        """
        sheet_name = "sortie"

        # Créer le chemin complet du fichier Excel
        fichier_sortie = os.path.join(dossier_data, 'sortie.xlsx')

        print(fichier_sortie)
        # Vérifier si le DataFrame n'est pas vide et si le fichier n'existe pas
        # if df is not None and not os.path.exists(fichier_sortie):
        if df is not None and not os.path.exists(fichier_sortie):
            # Créer un nouveau classeur
            workbook = Workbook()
            
            # Obtenir la feuille par défaut et la renommer
            feuille = workbook.active
            feuille.title = sheet_name

            # Écrire le DataFrame dans la première feuille
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    feuille.cell(row=r_idx, column=c_idx, value=value)

            # Supprimer la feuille par défaut s'il y en a une autre nommée 'Sheet'
            if 'Sheet' in workbook.sheetnames and 'Sheet' != sheet_name:
                default_sheet = workbook['Sheet']
                workbook.remove(default_sheet)

            # Sauvegarder le fichier Excel
            workbook.save(fichier_sortie)
            workbook.close()

            # Nettoyage mémoire
            gc.collect()

        return

    def step(self):
        """Effectue une étape de la simulation."""
        if self.t != 0:
            self.handle_poche()
            self.file_manager.etat_courant = self.handle_state(self.file_manager.etat_courant)
            self.record_data()
            Mgmin = self.file_manager.parametres_generaux['Mgmin']
            PFCmax = self.file_manager.parametres_generaux['PFCmax']
            PPT = self.file_manager.BoutonsEtParamètres['PPT']


            file_path = self.file_manager.button_file

            dossier_data = Path(file_path).resolve().parent
            df = self.construct_result_dataframe()
            self.export_result(df, dossier_data)



            

        self.t += 1

    def run_simulation(self, interval=1):
        """Démarre la simulation après initialisation."""
        self.file_manager.read_parametres_generaux()
        self.initialize_simulation()

        while self.file_manager.etat_courant != "etat_Fin":
            try:
                self.file_manager.check_for_updates()
                if self.file_manager.running:
                    self.step()
            except FileNotFoundError as e:
                print(f"Erreur : {e}")
            except Exception as e:
                print(f"Une erreur est survenue : {e}")
            time.sleep(interval)

    def handle_poche(self):
        """Gère les événements liés à la poche de magnésium."""
        if self.file_manager.DemandePoche:
            self.handle_message("DemandePoche")
            self.file_manager.DemandePoche = False
        if self.file_manager.AjoutPoche:
            self.Mg = self.file_manager.parametres_generaux["Mgmax"]
            self.PFC += self.file_manager.BoutonsEtParamètres["PPT"]
            self.handle_message("AjoutPoche")
            self.file_manager.AjoutPoche = False

    def handle_message(self, event):
        """Gère l'affichage des messages selon l'événement fourni."""
        if event == "DemandePoche":

            liste_quantites_moules_a_produire = self.file_manager.ProgrammeDISA["Nb Moules"]
            liste_masses_grappes_moules = self.file_manager.ProgrammeDISA["Poids"]
            liste_cadences_moule_par_heure = self.file_manager.ProgrammeDISA["Cadences"]
            

            R = self.file_manager.parametres_generaux['R']
            masse_fil = self.file_manager.parametres_generaux['masse_fil']
            masse_mg_fil = self.file_manager.parametres_generaux['masse_mg_fil']
            PFCmin = self.file_manager.parametres_generaux['PFCmin']
            Mgmin = self.file_manager.parametres_generaux['Mgmin']
            Mgmax = self.file_manager.parametres_generaux['Mgmax']
            eC = self.file_manager.parametres_generaux['eC']
            eP = self.file_manager.parametres_generaux['eP']
            temps_traitement = self.file_manager.parametres_generaux['temps_traitement']
            temps_gs = self.file_manager.parametres_generaux['temps_gs']
            temps_serie = self.file_manager.parametres_generaux['temps_serie']


            PPT = self.file_manager.BoutonsEtParamètres['PPT']
            S = self.file_manager.BoutonsEtParamètres['S']
            TPT = self.file_manager.BoutonsEtParamètres['TPT']



            delai_avt_traitement, Mglimite, PFClimite = calcul_temps_limite(self.PFC, self.Mg, PFCmin, Mgmin, eC, temps_traitement, temps_serie, liste_masses_grappes_moules, liste_cadences_moule_par_heure, liste_quantites_moules_a_produire)
            K, L = calcul_longueur_fil(PPT, S, temps_gs, TPT, R, masse_fil, 
                                       masse_mg_fil, Mgmax, eP, PFClimite, Mglimite)
            

            if delai_avt_traitement < 0 :
                 message = (f"Lancer la poche dans {delai_avt_traitement} minutes en prenant "
                       f"{L} m de fil fourée pour une visée de {K} en %")
            else :
                message = (f"Lancer la poche maintenant ({delai_avt_traitement} minutes de retard) en prenant "
                       f"avec {L} m de fil fourée pour une visée de {K} en %")
        elif event == "AjoutPoche":
            message = "On a ajouté la poche maintenant"


        data = {
            'Message': [message]
                }

        # Create the dataframe
        df = pd.DataFrame(data)
        return df


    def pendant_consommation(self, Mg, PFC):
        # Constantes dans ce programme
        eC = self.file_manager.parametres_generaux['eC']

        # Variables dans ce programme
        liste_quantites_moules_a_produire = self.file_manager.ProgrammeDISA["Nb Moules"]
        liste_masses_grappes_moules = self.file_manager.ProgrammeDISA["Poids"]
        liste_cadences_moule_par_heure = self.file_manager.ProgrammeDISA["Cadences"]

        global message

        # On produit les moules !! 
        masse_grappe_i = liste_masses_grappes_moules[0]
        cadence_moule_par_heure_i = liste_cadences_moule_par_heure[0]
        quantite_moules_a_produire_i = liste_quantites_moules_a_produire[0]
        
        # La cadence de consommation en kg/min du i-ème modèle
        cadence_fonte_i = cadence_moule_par_heure_i / 60 * masse_grappe_i  # en kg/min
        
        # Quantités de moules du i-ème modèle produits en une minute
        # cadence_moule_i = int(cadence_fonte_i / masse_grappe_i) # en unités/min
        cadence_moule_i = cadence_fonte_i / masse_grappe_i # en unités/min
        
        # Mise à jour du pourcentage de Mg et du poids fonte coulée
        PFC -= cadence_fonte_i
        Mg -= eC

        # On met à jour le programme de productions des moules
        liste_quantites_moules_a_produire[0] -= cadence_moule_i

        if liste_quantites_moules_a_produire[0] <= 0 :
            message ="On a théoriquement finis de produire ce modèle, passons au modèle suivant !"

        # On reste dans l'état consom tant que l'utilisateur ne dit pas de changer de Série
        etat_suivant= "etat_Consom"

        # Si on a fini de tout produire alors on stoppe la procédure 
        # On passe à l'état Fin
        if not liste_masses_grappes_moules:
            etat_suivant = "etat_Fin"
        
        return Mg, PFC, etat_suivant

    def pendant_changement_serie(self, Mg, PFC):
        # Constantes dans ce programme
        eC = self.file_manager.parametres_generaux['eC']
        
        # Variables dans ce programme
        temps_serie = self.file_manager.parametres_generaux['temps_serie']
        liste_quantites_moules_a_produire = self.file_manager.ProgrammeDISA["Nb Moules"]
        liste_masses_grappes_moules = self.file_manager.ProgrammeDISA["Poids"]
        liste_cadences_moule_par_heure = self.file_manager.ProgrammeDISA["Cadences"]


        temps_serie_courant = self.file_manager.parametres_generaux['temps_serie_courant']

        global message


        if temps_serie_courant == temps_serie :

            # Suppression des éléments déjà traités dans les listes
            liste_masses_grappes_moules.pop(0)
            liste_quantites_moules_a_produire.pop(0)
            liste_cadences_moule_par_heure.pop(0)
            
        # pendant Temps_Serie :
        Mg -= eC # perte de Mg par minute ou par seconde
        temps_serie_courant -= 1 # pas de temps en minute
       
        etat_suivant = "etat_Serie"


        if temps_serie_courant == 0:
            temps_serie_courant = temps_serie  # Réinitialiser le temps de série
            

            message ="On a finis le changement de modèle, commençons le prochain modèle !"

            etat_suivant = "etat_Consom"


        self.file_manager.parametres_generaux['temps_serie_courant'] = temps_serie_courant
        
        return Mg, PFC, etat_suivant

    def pendant_panne(self, Mg, PFC):
        # Variables dans ce programme
        Panne =  self.file_manager.Panne

        # Constantes dans ce programme
        eC = self.file_manager.parametres_generaux['eC']


        # pendant Temps_Panne non définis :
        Mg -= eC # perte de Mg par minute ou par seconde


        etat_suivant = "etat_Panne"
        if not Panne :
            etat_suivant = "etat_Consom"
        
        return Mg, PFC, etat_suivant

    def handle_state(self, etat_courant):
        """Gère les transitions d'états selon l'état courant."""
        if etat_courant == "etat_Consom":
            self.Mg, self.PFC, etat_suivant = self.pendant_consommation(self.Mg, self.PFC)
        elif etat_courant == "etat_Serie":
            self.Mg, self.PFC, etat_suivant = self.pendant_changement_serie(self.Mg, self.PFC)
        elif etat_courant == "etat_Panne":
            self.Mg, self.PFC, etat_suivant = self.pendant_panne(self.Mg, self.PFC)
        elif etat_courant == "etat_Fin":
            self.file_manager.running = False
            etat_suivant = "etat_Fin"
        return etat_suivant

    def record_data(self):
        """Enregistre les données de simulation pour suivi et affichage."""
        self.timedata.append(self.t)
        self.Mgdata.append(self.Mg)
        self.PFCdata.append(self.PFC)

    

