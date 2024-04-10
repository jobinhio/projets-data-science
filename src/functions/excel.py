import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows




def lire_fichiers_excel(chemin_fichier):
    # Lecture de la première feuille du fichier Excel : Table Matière Element
    df_table = pd.read_excel(chemin_fichier)
    # Lecture de la deuxième feuille du fichier Excel : Contraintes Element
    df_contraints = pd.read_excel(chemin_fichier, sheet_name=1)
    # Lecture de la troisième feuille du fichier Excel : Contraintes Matière Première
    df_MP = pd.read_excel(chemin_fichier, sheet_name=2)

    
    # Filtrage des articles disponibles et indisponibles
    df_MP_indispo = df_MP[df_MP['Disponible ?'] == 0].reset_index(drop=True)
    df_MP_dispo = df_MP[df_MP['Disponible ?'] == 1].reset_index(drop=True)

    # Separation des contraintes
    df_contraints_qualite = df_contraints.loc[:,['Composant','Impurété','ONO']]
    df_contraints_element = df_contraints.drop(columns=['Impurété', 'ONO'])

    return df_table,df_contraints_element, df_contraints_qualite, df_MP_dispo, df_MP_indispo



def write_dataframe_to_excel(df, excel_file_path, new_sheet_name='Recette'):
    """
    Écrit un DataFrame dans une nouvelle feuille Excel ou écrase une feuille existante avec les données du DataFrame.

    Args:
        df (DataFrame): Le DataFrame à écrire dans le fichier Excel.
        excel_file_path (str): Chemin du fichier Excel existant.
        new_sheet_name (str, optional): Nom de la nouvelle feuille Excel. Par défaut, 'Recette'.

    Returns:
        None
    """
    # Charger le classeur Excel
    workbook = load_workbook(excel_file_path)

    # Vérifier si la feuille existe déjà dans le classeur Excel
    if new_sheet_name in workbook.sheetnames:
        # Si la feuille existe déjà, effacer les données
        workbook.remove(workbook[new_sheet_name])

    # Ajouter la nouvelle feuille au classeur Excel
    workbook.create_sheet(title=new_sheet_name)

    # Charger la feuille nouvellement créée
    new_sheet = workbook[new_sheet_name]

    # Écrire les données dans la feuille
    for row in dataframe_to_rows(df, index=False, header=True):
        new_sheet.append(row)

    # Sauvegarder les modifications
    workbook.save(excel_file_path)

    # Fermer le classeur Excel
    workbook.close()